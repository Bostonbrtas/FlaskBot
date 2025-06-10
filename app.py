from quart import Quart, render_template, request, redirect, url_for, session, flash, current_app, send_file
from werkzeug.utils import secure_filename
from bson import ObjectId
from bson.errors import InvalidId
from motor.motor_asyncio import AsyncIOMotorClient

import os
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import logging
from datetime import datetime
from dateutil.parser import parse as parse_date
import webbrowser
import asyncio
from dateutil import parser as date_parser
from yadisk_utils import upload_to_yadisk, finalize_report
# Инициализация MongoDB
mongo_client = AsyncIOMotorClient("mongodb://localhost:27017")
db = mongo_client["users_db"]  # название базы можно изменить


# Допустимые расширения для фото
ALLOWED_PHOTO = {'png','jpg','jpeg','gif','pdf','dock'}
ALLOWED_SCAN  = ALLOWED_PHOTO

def _save_file(file, folder):
    os.makedirs(folder, exist_ok=True)
    filename = secure_filename(file.filename)
    path = os.path.join(folder, filename)
    file.save(path)
    return filename

# Список доступных должностей
POSITIONS = [
    "Монтажник",
    "Электрик",
    "Сантехник",
    "Прораб",
    "Инженер",
    "Другая"
]


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Путь к папке instance внутри server
INSTANCE_DIR = os.path.join(os.path.dirname(__file__), 'server', 'instance')
os.makedirs(INSTANCE_DIR, exist_ok=True)

# Инициализируем Flask с относительным instance_path
app = Quart(
    __name__,
    instance_relative_config=True,
    instance_path=INSTANCE_DIR,
    template_folder='server/templates',
    static_folder='static',
    static_url_path='/static'
)

# Конфигурация SQLAlchemy — единая БД в папке instance

app.secret_key = 'your_secret_key_here'

# Настройка загрузок
app.config['UPLOAD_FOLDER'] = os.path.join('static', 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# === Рест кода без изменений ===

@app.route('/fix_user_ids')
async def fix_user_ids():
    counter = 1
    async for user in db.users.find().sort("_id", 1):
        if "id" not in user:
            await db.users.update_one(
                {"_id": user["_id"]},
                {"$set": {"id": counter}}
            )
            counter += 1
    return "Готово"

@app.route("/fix_project_ids")
async def fix_project_ids():
    projects = await db.projects.find().to_list(None)

    # Сортируем по _id чтобы порядок был стабильным
    projects.sort(key=lambda x: str(x["_id"]))

    for idx, project in enumerate(projects, start=1):
        await db.projects.update_one(
            {"_id": project["_id"]},
            {"$set": {"id": idx}}
        )

    return "Проекты успешно пронумерованы."

@app.route('/check_report_times')
async def check_report_times():
    reports = await db.reports.find().to_list(5)  # Только 5 для примера
    for r in reports:
        print(f"Start: {r.get('start_time')}, End: {r.get('end_time')}")
    return "OK"

@app.route('/')
async def index():
    user_cursor = db.users.find({ "$or": [ {"archived": False}, {"archived": { "$exists": False }} ] })
    users = []
    async for user in user_cursor:
        user["id"] = str(user["_id"])
        users.append(user)

    return await render_template(
        'index.html',
        users=users,
        active_tab='users',
        archive=False
    )

@app.route('/login', methods=['GET', 'POST'])
async def login():
    if request.method == 'POST':
        data = await request.form
        username = data['username']
        password = data['password']

        if username == 'admin' and password == '1234':
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            return "Неверный логин или пароль", 403

    return await render_template('login.html')

@app.route('/edit_user/<user_id>', methods=['GET', 'POST'])
async def edit_user(user_id):
    try:
        _id = ObjectId(user_id)
    except InvalidId:
        return "Некорректный ID", 400

    user = await db.users.find_one({"_id": _id})
    if not user:
        return "Пользователь не найден", 404

    # числовой ID (например, 1, 2, 3)
    numeric_id = str(user.get("id"))

    if request.method == 'POST':
        form = await request.form
        files = await request.files

        # Обновляем основные поля
        update_fields = {
            "telegram_id": form['telegram_id'].strip(),
            "surname": form['surname'].strip(),
            "name": form['name'].strip(),
            "patronymic": form.get('patronymic') or None,
            "position": form.get('position'),
            "passport": form['passport'].strip(),
            "inn": form['inn'].strip(),
            "snils": form['snils'].strip(),
            "phone": form['phone'].strip(),
            "reg_address": form['reg_address'].strip(),
            "res_address": form.get('res_address') or None,
            "clothing_size": form.get('clothing_size') or None,
            "shoe_size": form.get('shoe_size') or None,
            "contract_type": form.get("contract_type")
        }

        bd_str = form.get('birth_date', '').strip()
        birth_date = None

        if bd_str:
            try:
                birth_date = datetime.strptime(bd_str, '%d.%m.%Y').date()
            except ValueError:
                try:
                    birth_date = datetime.strptime(bd_str, '%Y-%m-%d').date()
                except ValueError:
                    # Невалидная дата — логировать и проигнорировать
                    print("⚠️ Невалидная дата:", bd_str)
                    birth_date = None

        update_fields['birth_date'] = datetime.combine(birth_date, datetime.min.time()) if birth_date else None

        allowed_projects = form.getlist("allowed_projects")
        update_fields["allowed_projects"] = [ObjectId(pid) for pid in allowed_projects]

        # Фото
        photo = files.get('photo')
        if photo and photo.filename:
            photo_dir = os.path.join(app.static_folder, 'uploads', str(user.get("id")))
            os.makedirs(photo_dir, exist_ok=True)
            filename = f"photo_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{secure_filename(photo.filename)}"
            photo_path = os.path.join(photo_dir, filename)
            await photo.save(photo_path)
            update_fields["photo_path"] = f"uploads/{numeric_id}/{filename}"

        await db.users.update_one({"_id": _id}, {"$set": update_fields})

        # Дополнительные поля
        await db.user_fields.delete_many({"user_id": _id})
        names = form.getlist('field_name[]')
        values = form.getlist('field_value[]')
        for name, value in zip(names, values):
            if name.strip() and value.strip():
                await db.user_fields.insert_one({
                    "user_id": _id,
                    "field_name": name.strip(),
                    "field_value": value.strip()
                })

        # Сканы
        await db.user_scans.delete_many({"user_id": _id})
        scan_files = files.getlist('scan_file[]')
        scan_descs = form.getlist('scan_desc[]')
        scans_dir = os.path.join(app.static_folder, 'scans', str(user.get("id")))
        os.makedirs(scans_dir, exist_ok=True)

        for f, desc in zip(scan_files, scan_descs):
            if f and f.filename:
                fname = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{secure_filename(f.filename)}"
                fpath = os.path.join(scans_dir, fname)
                await f.save(fpath)
                await db.user_scans.insert_one({
                    "user_id": _id,
                    "scan_path": f"scans/{numeric_id}/{fname}",
                    "description": desc.strip(),
                    "filename": f.filename
                })

        return redirect('/')

    # GET-запрос
    user["allowed_projects"] = [ObjectId(pid) for pid in user.get("allowed_projects", [])]
    userfields = await db.user_fields.find({"user_id": _id}).to_list(None)
    scans = await db.user_scans.find({"user_id": _id}).to_list(None)
    projects = [p async for p in db.projects.find({"archived": False})]
    selected_project_ids = [str(pid) for pid in user.get("allowed_projects", [])]


    return await render_template(
        'edit_user.html',
        user=user,
        userfields=userfields,
        scans=scans,
        positions=POSITIONS,
        active_tab='users',
        projects=projects,
        selected_project_ids=selected_project_ids
    )

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg'}



@app.route('/add_user', methods=['GET', 'POST'])
async def add_user():
    if request.method == 'POST':
        form = await request.form
        files = await request.files

        tg_id = form['telegram_id'].strip()

        bd_str = form.get('birth_date', '').strip()
        birth_date = None

        if bd_str:
            try:
                birth_date = datetime.strptime(bd_str, '%d.%m.%Y').date()
            except ValueError:
                try:
                    birth_date = datetime.strptime(bd_str, '%Y-%m-%d').date()
                except ValueError:
                    print("⚠️ Невалидная дата:", bd_str)
                    birth_date = None


        # Проверка на дублирование
        existing_user = await db.users.find_one({"telegram_id": tg_id})
        if existing_user:
            return "Пользователь с таким Telegram ID уже существует.", 400

        allowed_project_ids = form.getlist("allowed_projects")
        allowed_projects = [ObjectId(pid) for pid in allowed_project_ids if pid]

        # Сбор полей
        user_data = {
            "telegram_id": tg_id,
            "surname": form['surname'].strip(),
            "name": form['name'].strip(),
            "patronymic": form.get('patronymic') or None,
            "position": form.get('position'),
            "passport": form['passport'].strip(),
            "inn": form['inn'].strip(),
            "snils": form['snils'].strip(),
            "phone": form['phone'].strip(),
            "reg_address": form['reg_address'].strip(),
            "res_address": form.get('res_address') or None,
            "clothing_size": form.get('clothing_size') or None,
            "shoe_size": form.get('shoe_size') or None,
            "birth_date": datetime.combine(birth_date, datetime.min.time()) if birth_date else None,
            "photo_path": None,
            "archived": False,
            "contract_type": form.get("contract_type"),
            "allowed_projects": allowed_projects,
        }

        # Вставка пользователя и получение его ID
        result = await db.users.insert_one(user_data)
        user_id = str(result.inserted_id)

        # === Фото ===
        photo = files.get('photo')
        if photo and photo.filename:
            photo_dir = os.path.join(app.static_folder, 'uploads', user_id)
            os.makedirs(photo_dir, exist_ok=True)
            filename = f"{datetime.utcnow().timestamp()}_{secure_filename(photo.filename)}"
            path = os.path.join(photo_dir, filename)
            await photo.save(path)
            photo_path = f"uploads/{user_id}/{filename}"
            await db.users.update_one(
                {"_id": ObjectId(user_id)},
                {"$set": {"photo_path": photo_path}}
            )

        # === Дополнительные поля ===
        names = form.getlist('field_name[]')
        values = form.getlist('field_value[]')
        for name, value in zip(names, values):
            if name.strip() and value.strip():
                await db.user_fields.insert_one({
                    "user_id": ObjectId(user_id),
                    "field_name": name.strip(),
                    "field_value": value.strip()
                })

        # === Сканы ===
        scan_files = files.getlist('scan_file')
        scan_descs = form.getlist('scan_desc')
        scan_dir = os.path.join(app.static_folder, 'scans', user_id)
        os.makedirs(scan_dir, exist_ok=True)

        for f, desc in zip(scan_files, scan_descs):
            if f and f.filename:
                fname = f"{datetime.utcnow().timestamp()}_{secure_filename(f.filename)}"
                fpath = os.path.join(scan_dir, fname)
                await f.save(fpath)
                await db.user_scans.insert_one({
                    "user_id": ObjectId(user_id),
                    "scan_path": f"scans/{user_id}/{fname}",
                    "description": desc.strip()
                })

        return redirect('/')
    projects = [p async for p in db.projects.find({"archived": False})]
    return await render_template('add_user.html', positions=POSITIONS, active_tab='users', projects=projects, selected_project_ids=[])

@app.route('/delete_user/<user_id>', methods=['POST'])
async def delete_user(user_id):
    try:
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            return "Пользователь не найден", 404

        # Удаление фото с диска
        if user.get("photo_path"):
            photo_path = os.path.join(app.static_folder, user["photo_path"])
            if os.path.exists(photo_path):
                os.remove(photo_path)

        # Удаление сканов пользователя
        async for scan in db.user_scans.find({"user_id": ObjectId(user_id)}):
            scan_path = os.path.join(app.static_folder, scan["scan_path"])
            if os.path.exists(scan_path):
                os.remove(scan_path)
        await db.user_scans.delete_many({"user_id": ObjectId(user_id)})

        # Удаление дополнительных полей
        await db.user_fields.delete_many({"user_id": ObjectId(user_id)})

        # Удаление самого пользователя
        await db.users.delete_one({"_id": ObjectId(user_id)})

        return redirect('/')

    except Exception as e:
        return await render_template('error.html', error=f"Ошибка при удалении пользователя: {e}")



@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/projects')
async def projects():
    sort = request.args.get('sort', 'alpha')
    direction = request.args.get('dir', 'asc')

    # Разрешённые сортировки
    valid_sorts = {
        'alpha': 'name',
        'address': 'address',
        'latitude': 'latitude',
        'longitude': 'longitude',
        'ask_location': 'ask_location'
    }
    sort_key = valid_sorts.get(sort, 'name')
    sort_order = 1 if direction == 'asc' else -1

    cursor = db.projects.find({"archived": False}).sort(sort_key, sort_order)
    projects = []
    async for p in cursor:
        p['id'] = str(p['_id'])  # добавляем id для шаблона
        projects.append(p)

    return await render_template(
        'projects.html',
        projects=projects,
        active_tab='projects',
        archive=False,
        sort=sort,
        dir=direction
    )

@app.route("/project/add", methods=["GET", "POST"])
async def add_project():
    if request.method == "POST":
        form = await request.form
        files = await request.files

        name = form["name"]
        address = form["address"]
        responsible_id = form["responsible_id"]
        ask_location = bool(form.get("ask_location"))

        try:
            latitude = float(form.get("latitude", ""))
            longitude = float(form.get("longitude", ""))
        except ValueError:
            await flash("Широта или долгота введены неверно", "danger")
            return redirect(url_for("add_project"))

        # Получение следующего числового ID
        last = await db.projects.find().sort("id", -1).limit(1).to_list(1)
        next_id = (last[0]["id"] + 1) if last else 1

        new_project = {
            "id": next_id,
            "name": name,
            "address": address,
            "latitude": latitude,
            "longitude": longitude,
            "responsible_id": responsible_id,
            "ask_location": ask_location,
            "archived": False
        }

        result = await db.projects.insert_one(new_project)
        project_id = str(result.inserted_id)

        scan_folder = os.path.join(current_app.static_folder, "scanProject", str(next_id))
        os.makedirs(scan_folder, exist_ok=True)

        for f in files.getlist("scans"):
            if f.filename:
                filename = secure_filename(f.filename)
                fpath = os.path.join(scan_folder, filename)
                await f.save(fpath)

                scan_doc = {
                    "project_id": project_id,
                    "scan_path": f"scanProject/{next_id}/{filename}",
                    "filename": filename
                }
                await db.project_scans.insert_one(scan_doc)

        return redirect(url_for("projects"))

    users_cursor = db.users.find()
    users = [user async for user in users_cursor]
    project = None
    return await render_template("add_project.html", users=users, project=project, active_tab='projects')


@app.route("/edit_project/<project_id>", methods=["GET", "POST"])
async def edit_project(project_id):
    project = await db.projects.find_one({"_id": ObjectId(project_id)})
    if not project:
        return "Проект не найден", 404

    numeric_id = str(project.get("id"))

    if request.method == "POST":
        form = await request.form
        files = await request.files

        updated_fields = {
            "name": form["name"],
            "address": form["address"],
            "ask_location": bool(form.get("ask_location", False))
        }

        responsible_id_str = form.get("responsible_id")
        if responsible_id_str:
            try:
                updated_fields["responsible_id"] = ObjectId(responsible_id_str)
            except Exception:
                pass

        try:
            updated_fields["latitude"] = float(form.get("latitude", ""))
            updated_fields["longitude"] = float(form.get("longitude", ""))
        except ValueError:
            pass

        await db.projects.update_one(
            {"_id": ObjectId(project_id)},
            {"$set": updated_fields}
        )

        delete_ids = form.getlist("delete_scans")
        for scan_id in delete_ids:
            scan = await db.project_scans.find_one({"_id": ObjectId(scan_id)})
            if scan:
                scan_path = os.path.join(app.static_folder, scan["scan_path"])
                if os.path.exists(scan_path):
                    os.remove(scan_path)
                await db.project_scans.delete_one({"_id": ObjectId(scan_id)})

        scan_folder = os.path.join(app.static_folder, "scanProject", numeric_id)
        os.makedirs(scan_folder, exist_ok=True)

        for f in files.getlist("scans"):
            if f and f.filename:
                filename = secure_filename(f.filename)
                path = os.path.join(scan_folder, filename)
                await f.save(path)

                await db.project_scans.insert_one({
                    "project_id": project_id,
                    "scan_path": f"scanProject/{numeric_id}/{filename}",
                    "filename": filename
                })

        return redirect(url_for("projects"))

    users = [u async for u in db.users.find()]
    scans = [s async for s in db.project_scans.find({"project_id": project_id})]

    return await render_template(
        "edit_project.html",
        project=project,
        users=users,
        scans=scans,
        active_tab='projects'
    )

@app.route('/reports')
async def show_reports():
    reports_cursor = db.reports.find({"archived": False}).sort("start_time", -1)
    reports = [r async for r in reports_cursor]

    # подтягиваем пользователей и проекты вручную
    for r in reports:
        r["user"] = await db.users.find_one({"_id": ObjectId(r["user_id"])}) if r.get("user_id") else None
        r["project"] = await db.projects.find_one({"_id": ObjectId(r["project_id"])}) if r.get("project_id") else None

    return await render_template(
        'reports.html',
        reports=reports,
        active_tab='reports',
        archive=False
    )


@app.route('/delete_project/<project_id>', methods=['POST'])
async def delete_project(project_id):
    try:
        project = await db.projects.find_one({"_id": ObjectId(project_id)})
        if not project:
            return "Проект не найден", 404

        # Удаляем все сканы проекта
        scans_cursor = db.project_scans.find({"project_id": project_id})
        async for scan in scans_cursor:
            scan_path = os.path.join(app.static_folder, scan["scan_path"])
            if os.path.exists(scan_path):
                os.remove(scan_path)
        await db.project_scans.delete_many({"project_id": project_id})

        # Удаляем сам проект
        await db.projects.delete_one({"_id": ObjectId(project_id)})

        logger.info(f"Проект с ID {project_id} удалён.")
        return redirect(url_for('projects'))

    except Exception as e:
        return await render_template('error.html', error=f"Ошибка при удалении проекта: {e}")


@app.route('/delete_project_scan/<scan_id>', methods=['POST'])
async def delete_project_scan(scan_id):
    scan = await db.project_scans.find_one({"_id": ObjectId(scan_id)})
    if not scan:
        return "Скан не найден", 404

    project_id = scan.get("project_id")

    # Удаление файла с диска
    scan_path = os.path.join(app.static_folder, scan["scan_path"])
    if os.path.exists(scan_path):
        os.remove(scan_path)

    # Удаление из базы
    await db.project_scans.delete_one({"_id": ObjectId(scan_id)})

    flash("Скан успешно удалён.")
    return redirect(url_for('edit_project', project_id=project_id))

@app.route('/delete_report/<report_id>', methods=['POST'])
async def delete_report(report_id):
    try:
        report = await db.reports.find_one({"_id": ObjectId(report_id)})
        if not report:
            return "Отчёт не найден", 404

        photo_dir = report.get("photo_dir")

        # Удаление всех связанных фотографий, если директория задана
        if photo_dir:
            full_dir = os.path.join(app.static_folder, photo_dir)
            if os.path.exists(full_dir):
                for filename in os.listdir(full_dir):
                    file_path = os.path.join(full_dir, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                os.rmdir(full_dir)

        # Удаление отчёта
        await db.reports.delete_one({"_id": ObjectId(report_id)})

        return redirect(url_for('show_reports'))

    except Exception as e:
        return await render_template('error.html', error=str(e))

@app.route('/export_users')
async def export_users():
    user_cursor = db.users.find()
    users = [user async for user in user_cursor]
    data = []

    for user in users:
        # Получаем дополнительные поля
        fields_cursor = db.user_fields.find({"user_id": ObjectId(user["_id"])})
        userfields = {f["field_name"]: f["field_value"] async for f in fields_cursor}

        data.append({
            'Telegram ID': user.get("telegram_id", ""),
            'Фамилия': user.get("surname", ""),
            'Имя': user.get("name", ""),
            'Отчество': user.get("patronymic", ""),
            'Дата рождения': user.get("birth_date", ""),
            'Должность': user.get("position", ""),
            'Паспорт': user.get("passport", ""),
            'ИНН': user.get("inn", ""),
            'СНИЛС': user.get("snils", ""),
            'Телефон': user.get("phone", ""),
            'Адрес прописки': user.get("reg_address", ""),
            'Адрес проживания': user.get("res_address", ""),
            'Размер одежды': user.get("clothing_size", ""),
            'Размер обуви': user.get("shoe_size", ""),
            **userfields
        })

    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)

    return await send_file(output, attachment_filename="сотрудники.xlsx", as_attachment=True)

@app.route('/export_projects')
async def export_projects():
    projects_cursor = db.projects.find()
    projects = [p async for p in projects_cursor]

    # Собираем данные
    data = []
    for p in projects:
        data.append({
            'ID': str(p["_id"]),
            'Название': p.get("name", ""),
            'Адрес': p.get("address", ""),
            'Широта': p.get("latitude", ""),
            'Долгота': p.get("longitude", "")
        })

    df = pd.DataFrame(data)
    file_path = os.path.join(app.instance_path, "projects_export.xlsx")
    df.to_excel(file_path, index=False)

    return await send_file(file_path, as_attachment=True)

@app.route('/export_reports')
async def export_reports():
    reports_cursor = db.reports.find()
    reports = [r async for r in reports_cursor]

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёты"

    headers = ["Telegram ID", "ФИО", "Проект", "Начало", "Конец", "Текст отчета", "Фото"]
    ws.append(headers)

    for report in reports:
        user = await db.users.find_one({"_id": ObjectId(report["user_id"])}) if report.get("user_id") else None
        project = await db.projects.find_one({"_id": ObjectId(report["project_id"])}) if report.get("project_id") else None

        telegram_id = user.get("telegram_id", "") if user else ""
        full_name = f"{user.get('surname', '')} {user.get('name', '')}".strip() if user else ""
        project_address = f"{project.get('name', '')}, {project.get('address', '')}".strip() if project else ""

        start_dt = report.get("start_time")
        end_dt = report.get("end_time")

        if isinstance(start_dt, datetime):
            start = start_dt.strftime('%Y-%m-%d %H:%M')
        else:
            start = ""

        if isinstance(end_dt, datetime):
            end = end_dt.strftime('%Y-%m-%d %H:%M')
        else:
            end = ""

        text = report.get("text_report", "")
        photo_link = report.get("photo_link", "")

        row = [telegram_id, full_name, project_address, start, end, text, ""]
        ws.append(row)

        if photo_link:
            cell = ws.cell(row=ws.max_row, column=7)
            cell.value = "Ссылка"
            cell.hyperlink = photo_link
            cell.font = Font(color="0000FF", underline="single")

    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    export_path = os.path.join(app.instance_path, "reports_export.xlsx")
    wb.save(export_path)

    return await send_file(export_path, attachment_filename="отчёты.xlsx", as_attachment=True)

@app.route("/sorted_reports")
async def sorted_reports():
    sort = request.args.get("sort", "start_time")
    order = request.args.get("order", "desc")

    sort_field_map = {
        "surname":     "user.surname",
        "telegram_id": "user.telegram_id",
        "project":     "project.name",
        "start_time":  "start_time",
        "end_time":    "end_time",
        "newest":      "_id",
    }

    reports_cursor = db.reports.find({"archived": {"$ne": True}})
    reports = []

    async for r in reports_cursor:
        user = await db.users.find_one({"_id": r["user_id"]}) if r.get("user_id") else None
        project = await db.projects.find_one({"_id": r["project_id"]}) if r.get("project_id") else None
        r["user"] = user
        r["project"] = project
        reports.append(r)

    sort_field = sort_field_map.get(sort, "start_time")

    def extract(obj, path):
        if path == "_id":
            return obj.get("_id")
        for part in path.split("."):
            obj = obj.get(part) if isinstance(obj, dict) else None
        return obj

    reverse = True if sort == "newest" else (order == "desc")

    def safe_sort_value(val):
        if isinstance(val, datetime):
            return val.isoformat()
        return str(val or "").lower()

    reports.sort(key=lambda r: safe_sort_value(extract(r, sort_field)), reverse=reverse)

    return await render_template(
        "reports.html",
        reports=reports,
        sort=sort,
        order=order,
        active_tab="reports",
        archive=False
    )

@app.route("/sorted_users")
async def sorted_users():
    sort = request.args.get("sort", "surname")
    order = request.args.get("order", "asc")

    sort_fields = ["surname", "name", "patronymic", "telegram_id", "position"]
    sort_field = sort if sort in sort_fields else "surname"
    reverse = (order == "desc")

    cursor = db.users.find({"archived": {"$ne": True}})
    users = [u async for u in cursor]

    def safe_val(user):
        val = user.get(sort_field)
        return str(val or "").lower()

    users.sort(key=safe_val, reverse=reverse)

    return await render_template(
        "index.html",
        users=users,
        sort=sort,
        order=order,
        active_tab="users",
        archive=False
    )

@app.route('/delete_user_scan/<scan_id>', methods=['POST'])
async def delete_user_scan(scan_id):
    try:
        _id = ObjectId(scan_id)
    except InvalidId:
        return "Некорректный ID", 400

    scan = await db.user_scans.find_one({"_id": _id})
    if not scan:
        return "Скан не найден", 404

    # Удаление файла с диска
    file_path = os.path.join(app.static_folder, scan["scan_path"])
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            print(f"⚠️ Ошибка при удалении файла: {e}")

    await db.user_scans.delete_one({"_id": _id})
    return '', 204

@app.route("/archive/users")
async def archive_users():
    sort_field = request.args.get("sort", "surname")
    order = request.args.get("order", "asc")
    reverse = order == "desc"

    # Ищем именно по archived, не по is_archived
    users_cursor = db.users.find({"archived": True})
    users = [u async for u in users_cursor]

    def extract(user, field):
        return str(user.get(field, "")).lower()

    users.sort(key=lambda u: extract(u, sort_field), reverse=reverse)

    return await render_template(
        "archive_users.html",
        users=users,
        active_tab="users",
        archive=True,
        sort=sort_field,
        order="desc" if not reverse else "asc"
    )

@app.route('/archive/projects')
async def archive_projects():
    sort = request.args.get('sort', 'alpha')
    direction = request.args.get('dir', 'asc')

    valid_sorts = {
        'alpha': 'name',
        'address': 'address',
        'latitude': 'latitude',
        'longitude': 'longitude',
        'ask_location': 'ask_location'
    }
    sort_field = valid_sorts.get(sort, 'name')

    cursor = db.projects.find({"archived": True}).sort(
        sort_field, -1 if direction == 'desc' else 1
    )
    projects = [p async for p in cursor]

    return await render_template(
        'archive_projects.html',
        projects=projects,
        active_tab='projects',
        archive=True,
        sort=sort,
        dir=direction
    )

@app.route('/archive/reports')
async def archive_reports():
    cursor = db.reports.find({"archived": True}).sort("start_time", -1)
    reports = []
    async for report in cursor:
        user = await db.users.find_one({"_id": ObjectId(report["user_id"])}) if report.get("user_id") else None
        project = await db.projects.find_one({"_id": ObjectId(report["project_id"])}) if report.get("project_id") else None
        report["user"] = user
        report["project"] = project
        reports.append(report)

    return await render_template(
        'archive_reports.html',
        reports=reports,
        active_tab='reports',
        archive=True
    )

@app.route("/sorted_archive_reports")
async def sorted_archive_reports():
    sort = request.args.get("sort", "start_time")
    order = request.args.get("order", "desc")

    sort_field_map = {
        "surname":     "user.surname",
        "telegram_id": "user.telegram_id",
        "project":     "project.name",
        "start_time":  "start_time",
        "end_time":    "end_time",
    }

    reports_cursor = db.reports.find({"archived": True})
    reports = []

    async for r in reports_cursor:
        user = await db.users.find_one({"_id": r["user_id"]}) if r.get("user_id") else None
        project = await db.projects.find_one({"_id": r["project_id"]}) if r.get("project_id") else None
        r["user"] = user
        r["project"] = project
        reports.append(r)

    sort_field = sort_field_map.get(sort, "start_time")

    def extract(obj, path):
        for part in path.split("."):
            obj = obj.get(part) if isinstance(obj, dict) else None
        return obj

    def safe(val):
        if isinstance(val, datetime):
            return val.isoformat()
        return str(val or "").lower()

    reports.sort(key=lambda r: safe(extract(r, sort_field)), reverse=(order == "desc"))

    return await render_template(
        "archive_reports.html",
        reports=reports,
        sort=sort,
        order=order,
        active_tab="reports",
        archive=True
    )

@app.route('/user/<user_id>/archive', methods=['POST'])
async def toggle_archive_user(user_id):
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        return "Пользователь не найден", 404

    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"archived": not user.get("archived", False)}}
    )
    return redirect(request.referrer or url_for('index'))

@app.route('/add_report', methods=['GET', 'POST'])
async def add_report():
    if request.method == 'POST':
        form = await request.form
        files = await request.files

        telegram_id = form.get('telegram_id')
        user_id = form.get('user_id')
        project_id = form.get('project_id')
        text_report = form.get('text_report', '').strip()

        try:
            start_time = date_parser.parse(form['start_time'], dayfirst=True)
            end_time = date_parser.parse(form['end_time'], dayfirst=True)
        except Exception:
            return "⛔ Неверный формат даты/времени. Пример: 2025-06-08 13:30 или 08.06.2025 13:30", 400

        uploaded_paths = []
        project = await db.projects.find_one({"_id": ObjectId(project_id)})
        user = await db.users.find_one({"_id": ObjectId(user_id)})

        for file in files.getlist("media_files"):
            if file and file.filename:
                file_bytes = file.read()
                ext = os.path.splitext(file.filename)[1].lower()
                filename = f"{datetime.now().strftime('%H-%M-%S-%f')}_{secure_filename(file.filename)}"
                last_folder = await upload_to_yadisk(db, project["name"], user["telegram_id"], file_bytes, filename)
                relpath = last_folder.removeprefix("Отчёты/") + "/" + filename
                uploaded_paths.append(relpath)

        # Сохраняем в базу
        report_data = {
            "user_id": ObjectId(user_id) if user_id else None,
            "project_id": ObjectId(project_id) if project_id else None,
            "telegram_id": telegram_id,  # ← добавь это
            "start_time": start_time,
            "end_time": end_time,
            "text_report": text_report,
            "archived": False
        }
        result = await db.reports.insert_one(report_data)
        report_id = result.inserted_id
        # Публикуем папку и получаем ссылку
        if uploaded_paths:
            public_url = finalize_report(last_folder)
            await db.reports.update_one(
                {"_id": report_id},
                {"$set": {"photo_link": public_url}}
            )

        for path in uploaded_paths:
            await db.report_photos.insert_one({
                "report_id": report_id,
                "photo_path": path
            })

        return redirect(url_for("show_reports"))

    # Преобразуем ObjectId → str, и ключ id вместо _id
    users_cursor = db.users.find({"archived": False})
    users = []
    async for user in users_cursor:
        users.append({
            "_id": str(user["_id"]),
            "telegram_id": user["telegram_id"],
            "full_name": f"{user.get('surname', '')} {user.get('name', '')}".strip()
        })

    projects_cursor = db.projects.find({"archived": False})
    projects = []
    async for project in projects_cursor:
        projects.append({
            "_id": str(project["_id"]),
            "name": project.get("name", "")
        })

    return await render_template("add_report.html", users=users, projects=projects)

@app.route('/project/<project_id>/archive', methods=['POST'])
async def toggle_archive_project(project_id):
    project = await db.projects.find_one({"_id": ObjectId(project_id)})
    if not project:
        return "Проект не найден", 404

    await db.projects.update_one(
        {"_id": ObjectId(project_id)},
        {"$set": {"archived": not project.get("archived", False)}}
    )
    return redirect(request.referrer or url_for('projects'))

@app.route('/report/<report_id>/archive', methods=['POST'])
async def toggle_archive_report(report_id):
    report = await db.reports.find_one({"_id": ObjectId(report_id)})
    if not report:
        return "Отчет не найден", 404

    await db.reports.update_one(
        {"_id": ObjectId(report_id)},
        {"$set": {"archived": not report.get("archived", False)}}
    )
    return redirect(request.referrer or url_for('show_reports'))


@app.route('/edit_report/<report_id>', methods=['GET', 'POST'])
async def edit_report(report_id):
    report = await db.reports.find_one({"_id": ObjectId(report_id)})
    if not report:
        return "Отчёт не найден", 404

    if request.method == 'POST':
        form = await request.form
        try:
            start_time = datetime.strptime(form['start_time'], '%d.%m.%Y, %H:%M')
        except ValueError:
            start_time = None

        end_str = form.get('end_time', '').strip()
        end_time = None
        if end_str:
            try:
                end_time = datetime.strptime(end_str, '%d.%m.%Y, %H:%M')
            except ValueError:
                pass

        await db.reports.update_one(
            {"_id": ObjectId(report_id)},
            {"$set": {
                "project_id": form['project_id'],
                "start_time": start_time,
                "end_time": end_time,
                "text_report": form['text_report'].strip()
            }}
        )
        return redirect(url_for('show_reports'))

    # GET
    projects_cursor = db.projects.find({"archived": False})
    projects = [p async for p in projects_cursor]

    return await render_template(
        'edit_report.html',
        report=report,
        projects=projects,
        active_tab='reports',
        archive=False
    )

async def open_browser():
    await asyncio.sleep(1)
    webbrowser.open_new("http://127.0.0.1:5001/")

if __name__ == '__main__':
    async def main():
        # Открытие браузера в фоне
        asyncio.create_task(open_browser())

        # Запуск приложения
        await app.run_task(host='0.0.0.0', port=5001, debug=True)

    asyncio.run(main())